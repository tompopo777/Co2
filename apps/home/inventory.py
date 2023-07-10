import datetime
import decimal
from urllib import request, parse, response
import pandas as pd
from IPython.core.display import display
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, Value, CharField
from django.db.models.functions import Cast
from decimal import *
from django.http import HttpResponse, FileResponse
from .count import *
import openpyxl
from .models import *
import json
from .views import carbon_system
from openpyxl import *
from openpyxl.utils.dataframe import *
from openpyxl.utils import *
from openpyxl.styles import *

pd.set_option('display.unicode.ambiguous_as_wide', True)
pd.set_option('display.unicode.east_asian_width', True)
pd.set_option('display.width', 200)


@login_required(login_url="/login/")
def inventory_summary(request):
    coefficient_source = request.session.get('coefficient_source')
    gwp_version = request.session.get('gwp_version')
    # 判斷使用者是否為公司帳號。
    if request.user.groups.filter(name='公司帳號').exists():
        factory_id = request.session.get('company_id')
    else:
        factory_id = request.session.get('factory_id')
    years = request.session.get('years')

    try:
        company_name = str(factory.objects.filter(id=factory_id).get())
        # print('company_name', company_name)
    except:
        company_name = ''

    # years = 2023
    # factory_id = 1
    # coefficient_source = '環保署溫室氣體排放係數管理表6.0.4'
    # gwp_version = 6
    emergency_generators_device = emergency_generators_inventory(years, factory_id, coefficient_source, gwp_version)
    combustion_equipment_device = combustion_equipment_inventory(years, factory_id, coefficient_source, gwp_version)
    official_car_device = official_car_inventory(years, factory_id, coefficient_source, gwp_version)
    refrigerant_device = refrigerant_inventory(years, factory_id, coefficient_source, gwp_version)
    personnel_inventory_device = personnel_inventory_inventory(years, factory_id, coefficient_source, gwp_version)
    employee_device = employee_inventory(years, factory_id, coefficient_source, gwp_version)
    # solvent_aerosol_emission_sources_device = solvent_aerosol_emission_sources_inventory(years, factory_id, coefficient_source, gwp_version)
    extinguisher_device = extinguisher_inventory(years, factory_id, coefficient_source, gwp_version)
    waste_water_device = waste_water_inventory(years, factory_id, coefficient_source, gwp_version)
    electricity_device = electricity_inventory(years, factory_id, coefficient_source, gwp_version)
    employee_commute_device = employee_commute_inventory(years, factory_id, coefficient_source, gwp_version)
    employee_business_trip_device = employee_business_trip_inventory(years, factory_id, coefficient_source, gwp_version)
    waste_transport_device = waste_transport_inventory(years, factory_id, coefficient_source, gwp_version)
    waste_process_device = waste_process_inventory(years, factory_id, coefficient_source, gwp_version)
    # purchase_material_device = purchase_material_inventory(years, factory_id, coefficient_source, gwp_version)
    category_one = pd.concat([emergency_generators_device, combustion_equipment_device, official_car_device, refrigerant_device, personnel_inventory_device,
                              employee_device, extinguisher_device, waste_water_device])
    category_two = pd.concat([electricity_device])
    category_three = pd.concat([employee_commute_device, employee_business_trip_device, waste_transport_device])
    category_four = pd.concat([waste_process_device])

    if category_one.empty and category_two.empty and category_three.empty and category_four:
        message = {
            'count_error': '沒有任何資料!'
        }
        request.method = "GET"
        return carbon_system(request, message)

    category_one['total_emission'] = category_one['emission'].sum()
    category_two['total_emission'] = category_two['emission'].sum()
    category_three['total_emission'] = category_three['emission'].sum()
    category_four['total_emission'] = category_four['emission'].sum()

    category_one = category_one.rename(columns={'process_area': '過程或區域', 'device_name': '排放源設施', 'fuel_type': '原燃物料', 'gas_name': '可能產生溫室氣體種類', 'emission': '排放當量公噸(公噸/數據期間)', 'total_emission': '加總排放當量(公噸CO2e/年)'})
    category_two = category_two.rename(columns={'process_area': '過程或區域', 'device_name': '排放源設施', 'fuel_type': '原燃物料', 'gas_name': '可能產生溫室氣體種類', 'emission': '排放當量公噸(公噸/數據期間)', 'total_emission': '加總排放當量(公噸CO2e/年)'})
    category_three = category_three.rename(columns={'process_area': '過程或區域', 'device_name': '排放源設施', 'fuel_type': '原燃物料', 'gas_name': '可能產生溫室氣體種類', 'emission': '排放當量公噸(公噸/數據期間)', 'total_emission': '加總排放當量(公噸CO2e/年)'})
    category_four = category_four.rename(columns={'process_area': '過程或區域', 'device_name': '排放源設施', 'fuel_type': '原燃物料', 'gas_name': '可能產生溫室氣體種類', 'emission': '排放當量公噸(公噸/數據期間)', 'total_emission': '加總排放當量(公噸CO2e/年)'})

    # 創建一個新的Excel工作簿
    workbook = Workbook()
    # 獲取預設的第一個工作表（預設情況下為"Sheet"）
    sheet = workbook.active
    # 將第一個工作表重新命名為您想要的名稱
    sheet.title = "溫室氣體排放清冊"

    # 設定標題欄位的樣式
    header_font = Font(bold=True)  # 設定粗體字體樣式
    border_style = Border(left=Side(style='thin', color='000000'),  # 設定左邊框線樣式為細線，顏色為黑色
                          right=Side(style='thin', color='000000'),  # 設定右邊框線樣式為細線，顏色為黑色
                          top=Side(style='thin', color='000000'),  # 設定頂部邊框線樣式為細線，顏色為黑色
                          bottom=Side(style='thin', color='000000'))  # 設定底部邊框線樣式為細線，顏色為黑色
    # 填滿顏色的樣式
    fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")  # 使用藍色作為填充顏色
    fill2 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    alignment = Alignment(horizontal='center', vertical='center')  # 設定水平居中對齊

    # 設置標題
    inventory_header = '溫室氣體排放總量'
    inventory_header_cell = sheet.cell(row=2, column=2, value=inventory_header)
    inventory_header_cell.font = header_font  # 設定額外標題的樣式
    inventory_header_cell.alignment = alignment  # 設定水平居中對齊
    # 合併欄位的儲存格
    merge_range = 'B2:G2'
    sheet.merge_cells(merge_range)
    # 設定合併儲存格的對齊方式為置中
    for row in sheet[merge_range]:
        for cell in row:
            cell.alignment = alignment

    # 將 DataFrame 資料輸出到 Excel 工作表指定位置
    rows_1 = dataframe_to_rows(category_one, index=False, header=True)
    start_row = 3  # 指定起始列
    start_column = 2  # 指定起始欄

    # 寫入標題欄位並設定樣式
    for c_idx, column in enumerate(category_one.columns, start=start_column):
        cell = sheet.cell(row=start_row, column=c_idx, value=column)
        cell.alignment = alignment  # 設定水平居中對齊
        cell.font = header_font  # 設定標題欄位的樣式
        cell.border = border_style  # 設定標題欄位的邊框樣式
        cell.fill = fill

    # 類別一開始--------------------------------------------------------------------------------------------------------------------------------
    # 插入一行字串
    category_one_header = ['類別一']  # 新增一行空白
    for c_idx, value in enumerate(category_one_header, start=start_column):
        cell = sheet.cell(row=start_row + 1, column=c_idx, value=value)
        cell.alignment = alignment  # 設定水平居中對齊
        cell.font = header_font  # 設定字串行的樣式
        cell.border = border_style
        cell.fill = fill2

    # 合併類別一標題欄的儲存格
    merge_range = f'B{start_row + 1}:G{start_row + 1}'
    sheet.merge_cells(merge_range)
    sheet[merge_range][0][0].alignment = alignment  # 設定合併儲存格的對齊方式為置中

    # 寫入數據並設定樣式
    for r_idx, row in enumerate(rows_1, start=start_row + 2):
        for c_idx, value in enumerate(row, start=start_column):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border_style

    # 合併 total_emission 欄位的儲存格
    start_row_merge = start_row + 3  # G4 的起始列
    end_row_merge = start_row_merge + len(category_one) - 1  # 合併到最後一筆資料的列
    merge_range = f'G{start_row_merge}:G{end_row_merge}'
    sheet.merge_cells(merge_range)

    # 設定合併儲存格的對齊方式為置中
    for row in sheet[merge_range]:
        for cell in row:
            cell.alignment = alignment

    # 類別二開始--------------------------------------------------------------------------------------------------------------------------------
    # 插入一行字串
    rows_2 = dataframe_to_rows(category_two, index=False, header=False)
    category_two_row = end_row_merge + 1
    category_two_header = ['類別二']  # 新增一行空白
    for c_idx, value in enumerate(category_two_header, start=start_column):
        cell = sheet.cell(row=category_two_row, column=c_idx, value=value)
        cell.alignment = alignment  # 設定水平居中對齊
        cell.font = header_font  # 設定字串行的樣式
        cell.border = border_style
        cell.fill = fill2

    # 合併類別一標題欄的儲存格
    merge_range = f'B{category_two_row}:G{category_two_row}'
    sheet.merge_cells(merge_range)
    sheet[merge_range][0][0].alignment = alignment  # 設定合併儲存格的對齊方式為置中

    # 寫入數據並設定樣式
    for r_idx, row in enumerate(rows_2, start=category_two_row + 1):
        for c_idx, value in enumerate(row, start=start_column):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border_style

    # 合併 total_emission 欄位的儲存格
    start_row2_merge = category_two_row + 1
    end_row2_merge = start_row2_merge + len(category_two) - 1  # 合併到最後一筆資料的列
    merge_range = f'G{start_row2_merge}:G{end_row2_merge}'
    sheet.merge_cells(merge_range)

    # 設定合併儲存格的對齊方式為置中
    for row in sheet[merge_range]:
        for cell in row:
            cell.alignment = alignment

    # 類別三開始--------------------------------------------------------------------------------------------------------------------------------
    # 插入一行字串
    rows_3 = dataframe_to_rows(category_three, index=False, header=False)
    category_three_row = end_row2_merge + 1
    category_three_header = ['類別三']  # 新增一行空白
    for c_idx, value in enumerate(category_three_header, start=start_column):
        cell = sheet.cell(row=category_three_row, column=c_idx, value=value)
        cell.alignment = alignment  # 設定水平居中對齊
        cell.font = header_font  # 設定字串行的樣式
        cell.border = border_style
        cell.fill = fill2

    # 合併類別一標題欄的儲存格
    merge_range = f'B{category_three_row}:G{category_three_row}'
    sheet.merge_cells(merge_range)
    sheet[merge_range][0][0].alignment = alignment  # 設定合併儲存格的對齊方式為置中

    # 寫入數據並設定樣式
    for r_idx, row in enumerate(rows_3, start=category_three_row + 1):
        for c_idx, value in enumerate(row, start=start_column):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border_style

    # 合併 total_emission 欄位的儲存格
    start_row3_merge = category_three_row + 1
    end_row3_merge = start_row3_merge + len(category_three) - 1  # 合併到最後一筆資料的列
    merge_range = f'G{start_row3_merge}:G{end_row3_merge}'
    sheet.merge_cells(merge_range)

    # 設定合併儲存格的對齊方式為置中
    for row in sheet[merge_range]:
        for cell in row:
            cell.alignment = alignment

    # 類別四開始--------------------------------------------------------------------------------------------------------------------------------
    # 插入一行字串
    rows_4 = dataframe_to_rows(category_two, index=False, header=False)
    category_four_row = end_row3_merge + 1
    category_four_header = ['類別四']  # 新增一行空白
    for c_idx, value in enumerate(category_four_header, start=start_column):
        cell = sheet.cell(row=category_four_row, column=c_idx, value=value)
        cell.alignment = alignment  # 設定水平居中對齊
        cell.font = header_font  # 設定字串行的樣式
        cell.border = border_style
        cell.fill = fill2

    # 合併類別一標題欄的儲存格
    merge_range = f'B{category_four_row}:G{category_four_row}'
    sheet.merge_cells(merge_range)
    sheet[merge_range][0][0].alignment = alignment  # 設定合併儲存格的對齊方式為置中

    # 寫入數據並設定樣式
    for r_idx, row in enumerate(rows_4, start=category_four_row + 1):
        for c_idx, value in enumerate(row, start=start_column):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border_style

    # 合併 total_emission 欄位的儲存格
    start_row4_merge = category_four_row + 1
    end_row4_merge = start_row4_merge + len(category_four) - 1  # 合併到最後一筆資料的列
    merge_range = f'G{start_row4_merge}:G{end_row4_merge}'
    sheet.merge_cells(merge_range)

    # 設定合併儲存格的對齊方式為置中
    for row in sheet[merge_range]:
        for cell in row:
            cell.alignment = alignment

    # 開始--------------------------------------------------------------------------------------------------------------------------------
    # 類別一七大氣體排放
    # 將 category_one_coefficient 資料框輸出到 Excel 的 K2 儲存格
    category_one_coefficient = category_one_coefficients(years, factory_id, coefficient_source, gwp_version)
    start_column_category = 9
    start_row_category = 4

    # 假設額外標題為 'Additional Header'
    category_one_header = '直接排放之七大溫室氣體排放量統計表'
    # 插入額外標題
    category_one_header_cell = sheet.cell(row=start_row_category - 1, column=start_column_category, value=category_one_header)
    category_one_header_cell.font = header_font  # 設定額外標題的樣式
    category_one_header_cell.alignment = alignment  # 設定水平居中對齊
    category_one_header_cell.fill = fill

    for c_idx, column in enumerate(category_one_coefficient.columns, start=start_column_category):
        cell = sheet.cell(row=start_row_category, column=c_idx, value=column)
        cell.alignment = Alignment(horizontal='center')  # 設定水平居中對齊
        cell.font = header_font  # 設定標題欄位的樣式
        cell.border = border_style  # 設定標題欄位的邊框樣式
        cell.fill = fill2

    for r_idx, row in enumerate(dataframe_to_rows(category_one_coefficient, index=False), start=start_row_category):
        for c_idx, value in enumerate(row, start=start_column_category):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')  # 設定水平居中對齊
            cell.border = border_style
            # 設定數字格式為保留四位小數
            cell.number_format = '0.0000'

    # 設定 Q5 儲存格的數值格式
    total_cell = sheet['Q5']
    total_cell.number_format = '0.000'

    # 合併欄位的儲存格
    merge_range = 'I3:Q3'
    sheet.merge_cells(merge_range)

    # 設定合併儲存格的對齊方式為置中
    for row in sheet[merge_range]:
        for cell in row:
            cell.alignment = alignment

    # 開始--------------------------------------------------------------------------------------------------------------------------------
    # 七大氣體排放
    # 將 all_coefficient 資料框輸出到 Excel 的 K2 儲存格
    all_coefficient = all_coefficients(years, factory_id, coefficient_source, gwp_version)
    start_column_all_category = 9
    start_row_all_category = 10

    # 假設額外標題為 'Additional Header'
    all_category_header = '溫室氣體排放量計算表'
    # 插入額外標題
    all_category_header_cell = sheet.cell(row=start_row_all_category - 1, column=start_column_all_category, value=all_category_header)
    all_category_header_cell.font = header_font  # 設定額外標題的樣式
    all_category_header_cell.alignment = alignment
    all_category_header_cell.fill = fill

    for c_idx, column in enumerate(all_coefficient.columns, start=start_column_all_category):
        cell = sheet.cell(row=start_row_all_category, column=c_idx, value=column)
        cell.alignment = Alignment(horizontal='center')  # 設定水平居中對齊
        cell.font = header_font  # 設定標題欄位的樣式
        cell.border = border_style  # 設定標題欄位的邊框樣式
        cell.fill = fill2

    for r_idx, row in enumerate(dataframe_to_rows(all_coefficient, index=False), start=start_row_all_category):
        for c_idx, value in enumerate(row, start=start_column_all_category):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')  # 設定水平居中對齊
            cell.border = border_style
            # 設定數字格式為保留四位小數
            cell.number_format = '0.0000'

    # 設定 Q5 儲存格的數值格式
    total_cell = sheet['Q11']
    total_cell.number_format = '0.000'

    # 合併欄位的儲存格
    merge_range = 'I9:Q9'
    sheet.merge_cells(merge_range)

    # 設定合併儲存格的對齊方式為置中
    merged_cells_range = sheet[merge_range]
    for cell_range in merged_cells_range:
        for cell in cell_range:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 開始--------------------------------------------------------------------------------------------------------------------------------
    # 各類排放係數
    all_category = all_categories(years, factory_id, coefficient_source, gwp_version)
    start_column_all_categories = 9
    start_row_all_categories = 16

    # 假設額外標題為 'Additional Header'
    all_categories_header = '溫室氣體排放量計算表'
    # 插入額外標題
    all_categories_header_cell = sheet.cell(row=start_row_all_categories - 1, column=start_column_all_categories, value=all_categories_header)
    all_categories_header_cell.font = header_font  # 設定額外標題的樣式
    all_categories_header_cell.alignment = alignment
    all_categories_header_cell.fill = fill

    for c_idx, column in enumerate(all_category.columns, start=start_column_all_categories):
        cell = sheet.cell(row=start_row_all_categories, column=c_idx, value=column)
        cell.alignment = Alignment(horizontal='center')  # 設定水平居中對齊
        cell.font = header_font  # 設定標題欄位的樣式
        cell.border = border_style  # 設定標題欄位的邊框樣式
        cell.fill = fill2

    for r_idx, row in enumerate(dataframe_to_rows(all_category, index=False), start=start_row_all_categories):
        for c_idx, value in enumerate(row, start=start_column_all_categories):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')  # 設定水平居中對齊
            cell.border = border_style
            # 設定數字格式為保留四位小數
            cell.number_format = '0.0000'

    # 設定 Q5 儲存格的數值格式
    total_cell = sheet['O17']
    total_cell.number_format = '0.000'

    # 合併欄位的儲存格
    merge_range = 'I15:O15'
    sheet.merge_cells(merge_range)

    # 設定合併儲存格的對齊方式為置中
    merged_cells_range = sheet[merge_range]
    for cell_range in merged_cells_range:
        for cell in cell_range:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 匯出檔名
    filename = '溫室氣體排放清冊-' + company_name + '_' + years + '.xlsx'

    # 返回 HttpResponse 響應給前端
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + parse.quote(filename, encoding="UTF-8")

    # 將工作簿內容寫入 HttpResponse
    workbook.save(response)

    return response


# 發電機
def emergency_generators_inventory(years, factory_id, coefficient_source, gwp_version):
    df = emergency_generators_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    # 合併所有的氣體
    row_data['gas_name'] = row_data['gas_name'].str.cat(sep=',')
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    # row_data = row_data.groupby(['process_area']).agg({'device_name': 'first', 'fuel_type': 'first', 'gas_name': 'first', 'emission': 'sum'}).reset_index()
    return row_data


# 燃燒設備
def combustion_equipment_inventory(years, factory_id, coefficient_source, gwp_version):
    df = combustion_equipment_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    # 合併所有的氣體
    row_data['fuel_type'] = row_data.groupby('device_name')['fuel_type'].transform(lambda x: ','.join(sorted(x)))
    row_data['gas_name'] = row_data.groupby('device_name')['gas_name'].transform(lambda x: ','.join(sorted(x)))
    # split將字串用逗號分割，然後再用set去除重複的
    row_data['fuel_type'] = row_data['fuel_type'].str.split(',').apply(lambda x: ','.join(set(x)))
    row_data['gas_name'] = row_data['gas_name'].str.split(',').apply(lambda x: ','.join(set(x)))
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 公務車
def official_car_inventory(years, factory_id, coefficient_source, gwp_version):
    df = official_car_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    # 合併所有的氣體
    row_data['fuel_type'] = row_data.groupby('device_name')['fuel_type'].transform(lambda x: ','.join(sorted(x)))
    row_data['gas_name'] = row_data.groupby('device_name')['gas_name'].transform(lambda x: ','.join(sorted(x)))
    # split將字串用逗號分割，然後再用set去除重複的
    row_data['fuel_type'] = row_data['fuel_type'].str.split(',').apply(lambda x: ','.join(set(x)))
    row_data['gas_name'] = row_data['gas_name'].str.split(',').apply(lambda x: ','.join(set(x)))
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 冷媒全部
def refrigerant_inventory(years, factory_id, coefficient_source, gwp_version):
    refrigerator = refrigerator_count(years, factory_id, coefficient_source, gwp_version)
    airconditioner = airconditioner_count(years, factory_id, coefficient_source, gwp_version)
    vehicle = vehicle_count(years, factory_id, coefficient_source, gwp_version)
    water_dispenser = water_dispenser_count(years, factory_id, coefficient_source, gwp_version)
    ice_water_dispenser = ice_water_dispenser_count(years, factory_id, coefficient_source, gwp_version)
    ice_maker = ice_maker_count(years, factory_id, coefficient_source, gwp_version)
    other_device = other_device_count(years, factory_id, coefficient_source, gwp_version)
    refrigerant_device = pd.concat([refrigerator, airconditioner, vehicle, water_dispenser, ice_water_dispenser, ice_maker, other_device], axis=0)
    row_data = refrigerant_device.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data['fuel_type'] = '冷媒'
    # 合併所有的氣體
    row_data['device_name'] = row_data.groupby('fuel_type')['device_name'].transform(lambda x: ','.join(sorted(x)))
    # # split將字串用逗號分割，然後再用set去除重複的
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 人天清冊
def personnel_inventory_inventory(years, factory_id, coefficient_source, gwp_version):
    df = personnel_inventory_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 委外人員
def employee_inventory(years, factory_id, coefficient_source, gwp_version):
    df = employee_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 滅火器
def extinguisher_inventory(years, factory_id, coefficient_source, gwp_version):
    df = extinguisher_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 厭氧廢水
def waste_water_inventory(years, factory_id, coefficient_source, gwp_version):
    df = waste_water_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 用電量
def electricity_inventory(years, factory_id, coefficient_source, gwp_version):
    df = electricity_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 員工通勤
def employee_commute_inventory(years, factory_id, coefficient_source, gwp_version):
    df = employee_commute_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data['device_name'] = '員工通勤'
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 員工出差
def employee_business_trip_inventory(years, factory_id, coefficient_source, gwp_version):
    df = employee_business_trip_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data['device_name'] = '員工出差'
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 廢棄物運輸
def waste_transport_inventory(years, factory_id, coefficient_source, gwp_version):
    df = waste_transport_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data['device_name'] = '廢棄物運輸'
    row_data['fuel_type'] = row_data['fuel_type'].str.split('(', 1).str[1].str.rstrip(')')
    row_data['fuel_type'] = row_data['fuel_type'].apply(lambda x: '運輸車輛-' + x)
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 廢棄物處理
def waste_process_inventory(years, factory_id, coefficient_source, gwp_version):
    df = waste_process_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 原物料採購
def purchase_material_inventory(years, factory_id, coefficient_source, gwp_version):
    df = purchase_material_count(years, factory_id, coefficient_source, gwp_version)
    row_data = df.drop(columns=['sum_count', 'data_unit', 'coefficient', 'coefficient_unit', 'coefficient_source', 'gwp_coefficient'])
    row_data = row_data.groupby(['process_area', 'device_name', 'fuel_type', 'gas_name'])['emission'].sum().reset_index()
    return row_data


# 類別一七大溫室氣體排放量統計
def category_one_coefficients(years, factory_id, coefficient_source, gwp_version):
    emergency_generators_device = emergency_generators_count(years, factory_id, coefficient_source, gwp_version)
    combustion_equipment_device = combustion_equipment_count(years, factory_id, coefficient_source, gwp_version)
    official_car_device = official_car_count(years, factory_id, coefficient_source, gwp_version)
    refrigerator_device = refrigerator_count(years, factory_id, coefficient_source, gwp_version)
    airconditioner_device = airconditioner_count(years, factory_id, coefficient_source, gwp_version)
    vehicle_device = vehicle_count(years, factory_id, coefficient_source, gwp_version)
    water_dispenser_device = water_dispenser_count(years, factory_id, coefficient_source, gwp_version)
    ice_water_dispenser_device = ice_water_dispenser_count(years, factory_id, coefficient_source, gwp_version)
    ice_maker_device = ice_maker_count(years, factory_id, coefficient_source, gwp_version)
    other_device_device = other_device_count(years, factory_id, coefficient_source, gwp_version)
    solvent_aerosol_emission_sources_device = solvent_aerosol_emission_sources_count(years, factory_id, coefficient_source, gwp_version)
    personnel_inventory_device = personnel_inventory_count(years, factory_id, coefficient_source, gwp_version)
    employee_device = employee_count(years, factory_id, coefficient_source, gwp_version)
    extinguisher_device = extinguisher_count(years, factory_id, coefficient_source, gwp_version)
    waste_water_device = waste_water_count(years, factory_id, coefficient_source, gwp_version)

    category_one = pd.concat([emergency_generators_device, combustion_equipment_device, official_car_device, refrigerator_device, airconditioner_device,
                              vehicle_device, water_dispenser_device, ice_water_dispenser_device, ice_maker_device, other_device_device, solvent_aerosol_emission_sources_device,
                              personnel_inventory_device, employee_device, extinguisher_device, waste_water_device])

    def coefficient_part(category_one):
        category_one_coefficient = pd.DataFrame(columns=['溫室氣體種類', 'CO2', 'CH4', 'N2O', 'HFCs', 'PFCs', 'SF6', 'NF3', '總量'])
        gas = '排放當量(公噸CO2e/年)'
        CO2 = Decimal(category_one.loc[category_one['gas_name'] == 'CO2', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        CH4 = Decimal(category_one.loc[category_one['gas_name'] == 'CH4', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        N2O = Decimal(category_one.loc[category_one['gas_name'] == 'N2O', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        HFCs = Decimal(category_one.loc[category_one['gas_name'] == 'HFCs', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        PFCs = Decimal(category_one.loc[category_one['gas_name'] == 'PFCs', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        SF6 = Decimal(category_one.loc[category_one['gas_name'] == 'SF6', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        NF3 = Decimal(category_one.loc[category_one['gas_name'] == 'NF3', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        total = Decimal(category_one.loc[category_one['gas_name'].isin(['CO2', 'CH4', 'N2O', 'HFCs', 'PFCs', 'SF6', 'NF3']), 'coefficient'].sum()).quantize(Decimal('0.000'), rounding=ROUND_HALF_UP)
        category_one_coefficient.loc[0] = [gas, CO2, CH4, N2O, HFCs, PFCs, SF6, NF3, total]
        return category_one_coefficient

    def percentages_part(category_one):
        category_one_percentages = coefficient_part(category_one)
        total_emission = category_one_percentages.loc[0, '總量']
        gas = '氣體別占比(％)'
        CO2 = str(Decimal(category_one_percentages.loc[0, 'CO2'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        CH4 = str(Decimal(category_one_percentages.loc[0, 'CH4'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        N2O = str(Decimal(category_one_percentages.loc[0, 'N2O'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        HFCs = str(Decimal(category_one_percentages.loc[0, 'HFCs'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        PFCs = str(Decimal(category_one_percentages.loc[0, 'PFCs'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        SF6 = str(Decimal(category_one_percentages.loc[0, 'SF6'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        NF3 = str(Decimal(category_one_percentages.loc[0, 'NF3'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        total = "100.00%"
        category_one_percentages.loc[0] = [gas, CO2, CH4, N2O, HFCs, PFCs, SF6, NF3, total]
        return category_one_percentages

    row1 = coefficient_part(category_one)
    row2 = percentages_part(category_one)
    final = pd.concat([row1, row2])

    return final


# 類別一七大溫室氣體排放量統計
def all_coefficients(years, factory_id, coefficient_source, gwp_version):
    emergency_generators_device = emergency_generators_count(years, factory_id, coefficient_source, gwp_version)
    combustion_equipment_device = combustion_equipment_count(years, factory_id, coefficient_source, gwp_version)
    official_car_device = official_car_count(years, factory_id, coefficient_source, gwp_version)
    refrigerator_device = refrigerator_count(years, factory_id, coefficient_source, gwp_version)
    airconditioner_device = airconditioner_count(years, factory_id, coefficient_source, gwp_version)
    vehicle_device = vehicle_count(years, factory_id, coefficient_source, gwp_version)
    water_dispenser_device = water_dispenser_count(years, factory_id, coefficient_source, gwp_version)
    ice_water_dispenser_device = ice_water_dispenser_count(years, factory_id, coefficient_source, gwp_version)
    ice_maker_device = ice_maker_count(years, factory_id, coefficient_source, gwp_version)
    other_device_device = other_device_count(years, factory_id, coefficient_source, gwp_version)
    solvent_aerosol_emission_sources_device = solvent_aerosol_emission_sources_count(years, factory_id, coefficient_source, gwp_version)
    personnel_inventory_device = personnel_inventory_count(years, factory_id, coefficient_source, gwp_version)
    employee_device = employee_count(years, factory_id, coefficient_source, gwp_version)
    extinguisher_device = extinguisher_count(years, factory_id, coefficient_source, gwp_version)
    waste_water_device = waste_water_count(years, factory_id, coefficient_source, gwp_version)
    electricity_device = electricity_count(years, factory_id, coefficient_source, gwp_version)
    employee_commute_device = employee_commute_count(years, factory_id, coefficient_source, gwp_version)
    employee_business_trip_device = employee_business_trip_count(years, factory_id, coefficient_source, gwp_version)
    waste_transport_device = waste_transport_count(years, factory_id, coefficient_source, gwp_version)
    waste_process_device = waste_process_count(years, factory_id, coefficient_source, gwp_version)
    purchase_material_device = purchase_material_count(years, factory_id, coefficient_source, gwp_version)

    category_all = pd.concat([emergency_generators_device, combustion_equipment_device, official_car_device, refrigerator_device, airconditioner_device,
                              vehicle_device, water_dispenser_device, ice_water_dispenser_device, ice_maker_device, other_device_device, solvent_aerosol_emission_sources_device,
                              personnel_inventory_device, employee_device, extinguisher_device, waste_water_device, electricity_device, employee_commute_device,
                              employee_business_trip_device, waste_transport_device, waste_process_device, purchase_material_device])

    def coefficient_part(category_all):
        all_coefficient = pd.DataFrame(columns=['溫室氣體種類', 'CO2', 'CH4', 'N2O', 'HFCs', 'PFCs', 'SF6', 'NF3', '總量'])
        gas = '排放當量(公噸CO2e/年)'
        CO2 = Decimal(category_all.loc[category_all['gas_name'] == 'CO2', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        CH4 = Decimal(category_all.loc[category_all['gas_name'] == 'CH4', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        N2O = Decimal(category_all.loc[category_all['gas_name'] == 'N2O', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        HFCs = Decimal(category_all.loc[category_all['gas_name'] == 'HFCs', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        PFCs = Decimal(category_all.loc[category_all['gas_name'] == 'PFCs', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        SF6 = Decimal(category_all.loc[category_all['gas_name'] == 'SF6', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        NF3 = Decimal(category_all.loc[category_all['gas_name'] == 'NF3', 'coefficient'].sum()).quantize(Decimal('0.0000'), rounding=ROUND_HALF_UP)
        total = Decimal(category_all.loc[category_all['gas_name'].isin(['CO2', 'CH4', 'N2O', 'HFCs', 'PFCs', 'SF6', 'NF3']), 'coefficient'].sum()).quantize(Decimal('0.000'), rounding=ROUND_HALF_UP)
        all_coefficient.loc[0] = [gas, CO2, CH4, N2O, HFCs, PFCs, SF6, NF3, total]
        return all_coefficient

    def percentages_part(category_all):
        all_percentages = coefficient_part(category_all)
        total_emission = all_percentages.loc[0, '總量']
        gas = '氣體別占比(％)'
        CO2 = str(Decimal(all_percentages.loc[0, 'CO2'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        CH4 = str(Decimal(all_percentages.loc[0, 'CH4'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        N2O = str(Decimal(all_percentages.loc[0, 'N2O'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        HFCs = str(Decimal(all_percentages.loc[0, 'HFCs'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        PFCs = str(Decimal(all_percentages.loc[0, 'PFCs'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        SF6 = str(Decimal(all_percentages.loc[0, 'SF6'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        NF3 = str(Decimal(all_percentages.loc[0, 'NF3'] / total_emission * 100).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)) + '%'
        total = "100.00%"
        all_percentages.loc[0] = [gas, CO2, CH4, N2O, HFCs, PFCs, SF6, NF3, total]
        return all_percentages

    row1 = coefficient_part(category_all)
    row2 = percentages_part(category_all)
    final = pd.concat([row1, row2])

    return final


# 各類排放係數
def all_categories(years, factory_id, coefficient_source, gwp_version):
    # years = 2023
    # factory_id = 1
    # coefficient_source = '環保署溫室氣體排放係數管理表6.0.4'
    # gwp_version = 6
    emergency_generators_device = emergency_generators_inventory(years, factory_id, coefficient_source, gwp_version)
    combustion_equipment_device = combustion_equipment_inventory(years, factory_id, coefficient_source, gwp_version)
    official_car_device = official_car_inventory(years, factory_id, coefficient_source, gwp_version)
    refrigerant_device = refrigerant_inventory(years, factory_id, coefficient_source, gwp_version)
    personnel_inventory_device = personnel_inventory_inventory(years, factory_id, coefficient_source, gwp_version)
    employee_device = employee_inventory(years, factory_id, coefficient_source, gwp_version)
    # solvent_aerosol_emission_sources_device = solvent_aerosol_emission_sources_inventory(years, factory_id, coefficient_source, gwp_version)
    extinguisher_device = extinguisher_inventory(years, factory_id, coefficient_source, gwp_version)
    waste_water_device = waste_water_inventory(years, factory_id, coefficient_source, gwp_version)
    electricity_device = electricity_inventory(years, factory_id, coefficient_source, gwp_version)
    employee_commute_device = employee_commute_inventory(years, factory_id, coefficient_source, gwp_version)
    employee_business_trip_device = employee_business_trip_inventory(years, factory_id, coefficient_source, gwp_version)
    waste_transport_device = waste_transport_inventory(years, factory_id, coefficient_source, gwp_version)
    waste_process_device = waste_process_inventory(years, factory_id, coefficient_source, gwp_version)
    # purchase_material_device = purchase_material_inventory(years, factory_id, coefficient_source, gwp_version)
    category_one = pd.concat([emergency_generators_device, combustion_equipment_device, official_car_device, refrigerant_device, personnel_inventory_device,
                              employee_device, extinguisher_device, waste_water_device])
    category_two = pd.concat([electricity_device])
    category_three = pd.concat([employee_commute_device, employee_business_trip_device, waste_transport_device])
    category_four = pd.concat([waste_process_device])

    all_category = pd.DataFrame(columns=['類別', '第1類', '第2類', '第3類', '第4類', '第5、6類', '總量'])
    category = '排放當量(公噸CO2e/年)'
    emission_category_1 = decimal.Decimal(category_one['emission'].sum().quantize(decimal.Decimal('0.0000'), rounding=decimal.ROUND_HALF_UP))
    emission_category_2 = decimal.Decimal(category_two['emission'].sum().quantize(decimal.Decimal('0.0000'), rounding=decimal.ROUND_HALF_UP))
    emission_category_3 = decimal.Decimal(category_three['emission'].sum().quantize(decimal.Decimal('0.0000'), rounding=decimal.ROUND_HALF_UP))
    emission_category_4 = decimal.Decimal(category_four['emission'].sum().quantize(decimal.Decimal('0.0000'), rounding=decimal.ROUND_HALF_UP))
    emission_category_5_6 = decimal.Decimal('0.0000')
    total_emission = (emission_category_1 + emission_category_2 + emission_category_3 + emission_category_4 + emission_category_5_6).quantize(decimal.Decimal('0.000'), rounding=decimal.ROUND_HALF_UP)
    all_category.loc[0] = [category, emission_category_1, emission_category_2, emission_category_3, emission_category_4, emission_category_5_6, total_emission]

    category2 = '氣體別占比(％)'
    percentage_category_1 = str((emission_category_1 / total_emission * 100).quantize(decimal.Decimal('0.00'))) + '%'
    percentage_category_2 = str((emission_category_2 / total_emission * 100).quantize(decimal.Decimal('0.00'))) + '%'
    percentage_category_3 = str((emission_category_3 / total_emission * 100).quantize(decimal.Decimal('0.00'))) + '%'
    percentage_category_4 = str((emission_category_4 / total_emission * 100).quantize(decimal.Decimal('0.00'))) + '%'
    percentage_category_5_6 = str((emission_category_5_6 / total_emission * 100).quantize(decimal.Decimal('0.00'))) + '%'
    total_percentage = '100%'
    all_category.loc[1] = [category2, percentage_category_1, percentage_category_2, percentage_category_3, percentage_category_4, percentage_category_5_6, total_percentage]

    return all_category
